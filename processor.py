from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from crf_reader import CrfReader
from excel_reader import ExcelReader
from json_generator import JsonGenerator
from models import AppConfig, Question, SurveyManifest
from xml_generator import XmlGenerator


class GiSTXProcessor:
    def __init__(self, config: AppConfig):
        self.config = config
        self.errorsEncountered = False
        self.logstring: list[str] = []
        self.generated_files: list[Path] = []
        self.question_list_cache: dict[str, list[Question]] = {}

    def run(self) -> int:
        self.logstring = [f"Log file for: {self.config.excelFile}"]

        excel_path = Path(self.config.excelFile)
        output_path = Path(self.config.outputPath)
        output_path.mkdir(parents=True, exist_ok=True)

        if not excel_path.exists():
            self.logstring.append("ERROR: Excel file not found!")
            self.logstring.append(str(excel_path))
            self.logstring.extend(
                [
                    "\r--------------------------------------------------------------------------------",
                    "End of log file",
                    "--------------------------------------------------------------------------------",
                ]
            )
            self._write_logfile()
            return 1

        workbook = load_workbook(filename=excel_path, data_only=False)

        try:
            worksheets = [ws for ws in workbook.worksheets if ws.title.endswith("_dd") or ws.title.endswith("_xml")]

            for ws in worksheets:
                reader = ExcelReader()
                qlist = reader.create_question_list(ws)
                if reader.errorsEncountered:
                    self.errorsEncountered = True
                self.logstring.extend(reader.logstring)
                self.question_list_cache[ws.title] = qlist

            xml_files: list[str] = []

            if not self.errorsEncountered:
                for ws_name, qlist in self.question_list_cache.items():
                    xml_name = ws_name.replace("_dd", ".xml").replace("_xml", ".xml")
                    xml_files.append(xml_name)

                    xml_generator = XmlGenerator()
                    xml_path = xml_generator.write_xml(ws_name, qlist, output_path)
                    self.logstring.extend(xml_generator.logstring)

                    if not self._validate_xml_syntax(xml_path):
                        self.errorsEncountered = True
                    self.generated_files.append(xml_path)

                crfs = []
                crfs_ws = workbook["crfs"] if "crfs" in workbook.sheetnames else None
                if crfs_ws is not None:
                    crfs = CrfReader.read_crfs_worksheet(crfs_ws)

                manifest = SurveyManifest(
                    surveyName=self.config.surveyName,
                    surveyId=self.config.surveyId,
                    databaseName=f"{self.config.surveyId}.sqlite",
                    xmlFiles=xml_files,
                    crfs=crfs,
                )
                manifest_path = output_path / "survey_manifest.gistx"
                JsonGenerator.write_manifest(manifest_path, manifest)
                self.logstring.append("")
                self.logstring.append("Successfully generated survey_manifest.gistx")
                self.generated_files.append(manifest_path)

            self.logstring.extend(
                [
                    "\r--------------------------------------------------------------------------------",
                    "End of log file",
                    "--------------------------------------------------------------------------------",
                ]
            )

            if not self.errorsEncountered:
                self._create_zip_file()

            self._write_logfile()
            return 1 if self.errorsEncountered else 0
        finally:
            workbook.close()

    def _validate_xml_syntax(self, file_path: Path) -> bool:
        try:
            ET.parse(file_path)
            return True
        except ET.ParseError as ex:
            self.logstring.append(f"CRITICAL ERROR: XML Syntax Error in file '{file_path.name}'")
            self.logstring.append(f"Details: {ex}")
            return False
        except Exception as ex:
            self.logstring.append(f"CRITICAL ERROR: Could not validate XML file '{file_path.name}'")
            self.logstring.append(f"Details: {ex}")
            return False

    def _write_logfile(self) -> None:
        logfile = Path(self.config.outputPath) / "gistlogfile.txt"
        with logfile.open("w", encoding="utf-8", newline="\n") as f:
            for line in self.logstring:
                f.write(line + "\n")
            f.write("\n")

    def _create_zip_file(self) -> None:
        zip_file_path = Path(self.config.outputPath) / f"{self.config.surveyId}.zip"
        if zip_file_path.exists():
            zip_file_path.unlink()

        with ZipFile(zip_file_path, "w", compression=ZIP_DEFLATED) as archive:
            for file_path in self.generated_files:
                if file_path.exists():
                    archive.write(file_path, arcname=file_path.name)
                    self.logstring.append(f"Added to zip: {file_path.name}")

            if self.config.csvFiles:
                csv_dir = Path(self.config.csvFiles.rstrip("\\/"))
                if csv_dir.exists() and csv_dir.is_dir():
                    csv_files = sorted(csv_dir.glob("*.csv"))
                    if csv_files:
                        self.logstring.append("")
                        self.logstring.append("Adding CSV files to package:")
                        for csv in csv_files:
                            archive.write(csv, arcname=csv.name)
                            self.logstring.append(f"Added to zip: {csv.name}")
                    else:
                        self.logstring.append(f"WARNING: No CSV files found in {csv_dir}")
                else:
                    self.logstring.append(f"WARNING: CSV files directory not found: {csv_dir}")

        self.logstring.append("")
        self.logstring.append(f"Successfully created zip file: {zip_file_path}")

        for file_path in self.generated_files:
            if file_path.exists():
                file_path.unlink()
                self.logstring.append(f"Deleted temporary file: {file_path.name}")


def run_from_config_file(config_file: str | Path) -> int:
    cfg_path = Path(config_file)
    data = json.loads(cfg_path.read_text(encoding="utf-8"))
    config = AppConfig(
        excelFile=data.get("excelFile", ""),
        csvFiles=data.get("csvFiles", ""),
        outputPath=data.get("outputPath", ""),
        surveyName=data.get("surveyName", ""),
        surveyId=data.get("surveyId", ""),
    )
    processor = GiSTXProcessor(config)
    return processor.run()
