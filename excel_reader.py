from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from models import (
    CalculationParameter,
    CalculationPart,
    CalculationType,
    CaseCondition,
    Filter,
    Question,
    ResponseSourceType,
)

class ExcelReader:
    NUMBER_OF_COLUMNS = 14
    COLUMN_NAMES = [
        "FieldName",
        "QuestionType",
        "FieldType",
        "QuestionText",
        "MaxCharacters",
        "Responses",
        "LowerRange",
        "UpperRange",
        "LogicCheck",
        "DontKnow",
        "Refuse",
        "NA",
        "Skip",
        "Comments",
    ]

    NUMERIC_ONLY_RE = re.compile(r"^\d+$")
    DECIMAL_RE = re.compile(r"^\d+(\.\d+)?$")
    DATE_RANGE_RE = re.compile(r"^([+-])(\d+)([dwmy])$")
    HARDCODED_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    FIELD_NAME_RE = re.compile(r"\b[a-z_][a-z0-9_]*\b", re.IGNORECASE)
    QUOTED_STRING_RE = re.compile(r"'[^']*'")
    FILTER_MATCH_RE = re.compile(r"^(\w+)\s*(?:(=|!=|<>|>|<|>=|<=)\s*)?(.+)$")
    PARAMETER_RE = re.compile(r"^(@?\w+)\s*=\s*(\w+)$")
    WHEN_CONDITION_RE = re.compile(r"^(\w+)\s+(=|!=|<>|>=|<=|>|<)\s+(.+?)\s*=>\s*(.+)$")

    VALID_QUESTION_TYPES = {
        "radio",
        "combobox",
        "checkbox",
        "text",
        "date",
        "information",
        "automatic",
        "button",
    }
    VALID_FIELD_TYPES = {
        "text",
        "datetime",
        "date",
        "phone_num",
        "integer",
        "text_integer",
        "text_decimal",
        "text_id",
        "n/a",
        "hourmin",
    }
    BUILT_IN_AUTO_FIELDS = {"starttime", "stoptime", "uniqueid", "swver", "survey_id", "lastmod"}
    LOGIC_KEYWORDS = {"and", "or", "not"}

    def __init__(self) -> None:
        self.logstring: list[str] = []
        self.errorsEncountered = False
        self.worksheetErrorsEncountered = False
        self.questionList: list[Question] = []

    def create_question_list(self, worksheet: Worksheet) -> list[Question]:
        self.worksheetErrorsEncountered = False
        self.logstring.append(f"\rChecking worksheet: '{worksheet.title}'")
        self.questionList = []

        for row_idx in range(1, worksheet.max_row + 1):
            try:
                if row_idx == 1:
                    current_headers = [self._get_cell_trim(worksheet, 1, i + 1) for i in range(self.NUMBER_OF_COLUMNS)]
                    if current_headers != self.COLUMN_NAMES:
                        self._error(
                            "ERROR: The header names in the "
                            f"{worksheet.title} are incorrect. Header names should be: "
                            "FieldName, QuestionType, FieldType, QuestionText, MaxCharacters, "
                            "Responses, LowerRange, UpperRange, LogicCheck, DontKnow, Refuse, NA, Skip, Comments"
                        )
                    continue

                if self._is_cell_merged(worksheet, row_idx, self.NUMBER_OF_COLUMNS):
                    continue

                q = Question()
                q.fieldName = self._get_cell_trim(worksheet, row_idx, 1)
                if not q.fieldName:
                    self._error(
                        f"ERROR - FieldName: Row {row_idx} in worksheet '{worksheet.title}' has a blank FieldName."
                    )
                    continue

                self._check_field_name(worksheet.title, q.fieldName)
                q.questionType = self._get_cell_trim(worksheet, row_idx, 2)
                q.fieldType = self._get_cell_trim(worksheet, row_idx, 3)
                q.questionText = self._get_cell_trim(worksheet, row_idx, 4)

                if q.questionText == "" and q.questionType != "automatic":
                    self._error(
                        f"ERROR - QuestionText: FieldName '{q.fieldName}' in worksheet '{worksheet.title}' has blank QuestionText."
                    )

                max_chars = self._get_cell_trim(worksheet, row_idx, 5)
                q.maxCharacters = max_chars if max_chars else "-9"
                if q.maxCharacters != "-9":
                    self._check_max_chars_value(worksheet.title, q.maxCharacters, q.fieldName)

                raw_responses = self._get_cell_raw(worksheet, row_idx, 6)
                raw_stripped = raw_responses.strip()
                if raw_stripped.lower().startswith("source:"):
                    self._parse_dynamic_responses(raw_responses, q, worksheet.title, q.fieldName)
                elif raw_stripped.lower().startswith("calc:"):
                    if q.questionType == "automatic":
                        if q.fieldName.lower() not in self.BUILT_IN_AUTO_FIELDS:
                            self._parse_automatic_calculation(raw_responses, q, worksheet.title, q.fieldName)
                    else:
                        self._error(
                            f"ERROR - Calculation: FieldName '{q.fieldName}' in worksheet '{worksheet.title}' "
                            "has calculation syntax but QuestionType is not 'automatic'."
                        )
                elif raw_stripped.lower().startswith("mask:"):
                    if q.questionType == "text":
                        q.mask = raw_stripped[5:].strip()
                    else:
                        self._error(
                            f"ERROR - Mask: FieldName '{q.fieldName}' in worksheet '{worksheet.title}' "
                            "has mask syntax but QuestionType is not 'text'."
                        )
                else:
                    q.responses = raw_responses

                self._check_question_field_type(q, worksheet.title)

                lower_val = self._get_cell_trim(worksheet, row_idx, 7)
                upper_val = self._get_cell_trim(worksheet, row_idx, 8)
                q.lowerRange = lower_val if lower_val else "-9"
                q.upperRange = upper_val if upper_val else "-9"

                if q.questionType == "date":
                    self._check_date_range(worksheet.title, q.lowerRange, q.fieldName, "LowerRange")
                    self._check_date_range(worksheet.title, q.upperRange, q.fieldName, "UpperRange")
                else:
                    if q.lowerRange != "-9":
                        self._check_numeric_range(worksheet.title, q.lowerRange, q.fieldName, "LowerRange")
                    if q.upperRange != "-9":
                        self._check_numeric_range(worksheet.title, q.upperRange, q.fieldName, "UpperRange")

                logic_raw = self._get_cell_trim(worksheet, row_idx, 9)
                if logic_raw:
                    for check in self._split_lines(logic_raw):
                        trimmed = check.strip()
                        if trimmed.startswith("unique;"):
                            parts = trimmed.split(";", 1)
                            if len(parts) == 2:
                                message = parts[1].strip()
                                if message.startswith("'") and message.endswith("'"):
                                    q.uniqueCheckMessage = message.strip("'")
                                else:
                                    self._error(
                                        f"ERROR - LogicCheck: FieldName '{q.fieldName}' in worksheet '{worksheet.title}' "
                                        "has invalid syntax for unique check message (must be in single quotes): "
                                        f"{trimmed}"
                                    )
                            else:
                                self._error(
                                    f"ERROR - LogicCheck: FieldName '{q.fieldName}' in worksheet '{worksheet.title}' "
                                    f"has invalid syntax for unique check (missing message): {trimmed}"
                                )
                        else:
                            q.logicChecks.append(trimmed)
                            self._check_logic_check_syntax(worksheet.title, trimmed, q.fieldName)

                q.dontKnow = self._get_cell_trim(worksheet, row_idx, 10) or "-9"
                if q.dontKnow != "-9":
                    self._check_special_button(worksheet.title, q.dontKnow, q.fieldName, "DontKnow")

                q.refuse = self._get_cell_trim(worksheet, row_idx, 11) or "-9"
                if q.refuse != "-9":
                    self._check_special_button(worksheet.title, q.refuse, q.fieldName, "Refuse")

                q.na = self._get_cell_trim(worksheet, row_idx, 12) or "-9"
                if q.na != "-9":
                    self._check_special_button(worksheet.title, q.na, q.fieldName, "NA")

                q.skip = self._get_cell_trim(worksheet, row_idx, 13)
                if q.skip:
                    self._check_skip_syntax(worksheet.title, q.skip, q.fieldName)

                self.questionList.append(q)
            except Exception as ex:
                self._error(
                    f"ERROR: An unexpected error occurred while processing row {row_idx} in worksheet "
                    f"'{worksheet.title}'. The error was: {ex}"
                )

        if not self.worksheetErrorsEncountered:
            self._check_logic_field_names(worksheet.title)
            self._check_skip_to_field_names(worksheet.title)
            self._check_required_max_characters(worksheet.title)
            self._check_duplicate_columns(worksheet.title)
            if not self.worksheetErrorsEncountered:
                self.logstring.append(f"No errors found in '{worksheet.title}'")

        return self.questionList

    @staticmethod
    def count_data_rows(worksheet: Worksheet) -> int:
        count = 0
        for row_idx in range(2, worksheet.max_row + 1):
            if not ExcelReader._is_cell_merged(worksheet, row_idx, 14):
                count += 1
        return count

    @staticmethod
    def _split_lines(text: str) -> list[str]:
        return [line for line in re.split(r"\r\n|\n|\r", text) if line]

    @staticmethod
    def _to_str(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    def _get_cell_trim(self, ws: Worksheet, row: int, col: int) -> str:
        return self._to_str(ws.cell(row=row, column=col).value).strip()

    def _get_cell_raw(self, ws: Worksheet, row: int, col: int) -> str:
        return self._to_str(ws.cell(row=row, column=col).value)

    @staticmethod
    def _is_cell_merged(ws: Worksheet, row: int, col: int) -> bool:
        coord = ws.cell(row=row, column=col).coordinate
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                return True
        return False

    def _error(self, message: str) -> None:
        self.errorsEncountered = True
        self.worksheetErrorsEncountered = True
        self.logstring.append(message)

    def _check_field_name(self, worksheet: str, fieldname: str) -> None:
        if not fieldname:
            self._error(f"ERROR - FieldName: {worksheet} has an empty FieldName")
            return
        if fieldname[0].isdigit():
            self._error(f"ERROR - FieldName: {worksheet} has a FieldName that starts with a number: {fieldname}")
        elif any((not c.isalnum()) and c != "_" for c in fieldname):
            self._error(
                "ERROR - FieldName: "
                f"{worksheet} has an invalid FieldName.  Only letters, digits, and underscores are allowed: {fieldname}"
            )
        elif fieldname != fieldname.lower():
            self._error(f"ERROR - FieldName: {worksheet} has a FieldName that is not all lowercase: {fieldname}")
        elif fieldname[0] == "_":
            self._error(f"ERROR - FieldName: {worksheet} has a FieldName that starts with an underscore: {fieldname}")
        elif " " in fieldname:
            self._error(f"ERROR - FieldName: {worksheet} has a FieldName that contains a space: {fieldname}")

    def _check_max_chars_value(self, worksheet: str, max_chars: str, fieldname: str) -> None:
        numeric = max_chars[1:] if max_chars.startswith("=") else max_chars
        if not self.NUMERIC_ONLY_RE.fullmatch(numeric):
            self._error(
                f"ERROR - MaxCharacters: FieldName '{fieldname}' in worksheet '{worksheet}' "
                f"has a non-numeric value for MaxCharacters: {max_chars}"
            )
            return
        num = int(numeric)
        if num < 1 or num > 2000:
            self._error(
                f"ERROR - MaxCharacters: FieldName '{fieldname}' in worksheet '{worksheet}' "
                f"has a MaxCharacters value that is out of range (1 to 2000): {max_chars}"
            )

    def _check_question_field_type(self, q: Question, worksheet: str) -> None:
        questiontype = q.questionType
        fieldtype = q.fieldType
        fieldname = q.fieldName

        if questiontype not in self.VALID_QUESTION_TYPES:
            self._error(
                f"ERROR - QuestionType: The QuestionType {questiontype} for FieldName '{fieldname}' "
                f"in table '{worksheet}' is not among the predefined list."
            )

        if fieldtype not in self.VALID_FIELD_TYPES:
            self._error(
                f"ERROR - FieldType: The FieldType '{fieldtype}' for FieldName '{fieldname}' in table '{worksheet}' "
                "is not among the predefined list."
            )

        if questiontype == "radio" and fieldtype != "integer":
            self._error(
                f"ERROR - FieldType: The FieldType for FieldName '{fieldname}' in table '{worksheet}' must be integer "
                "when the QuestionType is 'radio'."
            )

        if questiontype == "checkbox" and fieldtype != "text":
            self._error(
                f"ERROR - FieldType: The FieldType for FieldName '{fieldname}' in table '{worksheet}' must be text "
                "when the QuestionType is 'checkbox'."
            )

        if questiontype == "date" and fieldtype not in {"date", "datetime"}:
            self._error(
                f"ERROR - FieldType: The FieldType for FieldName '{fieldname}' in table '{worksheet}' "
                "must be date when the QuestionType is 'date' or 'datetime'."
            )

        if questiontype in {"radio", "checkbox"} and q.responseSourceType == ResponseSourceType.STATIC and q.responses:
            responses = self._split_lines(q.responses)
            seen: list[str] = []
            for response in responses:
                index = response.find(":")
                if index == -1:
                    self._error(
                        f"ERROR - Responses: Invalid static radio button options for '{fieldname}' in table '{worksheet}'. "
                        f"Expected format 'number:Statement', found '{response}'."
                    )
                    return
                if len(response.split(":")) != 2:
                    self._error(
                        f"ERROR - Responses: Invalid static radio button options for '{fieldname}' in table '{worksheet}'. "
                        f"Expected format 'number:Statement', found '{response}'."
                    )
                    return
                key = response[:index]
                seen.append(key)
                duplicates = sorted({k for k in seen if seen.count(k) > 1})
                if len(set(seen)) != len(seen):
                    self._error(
                        f"ERROR - Responses: The Responses for FieldName '{fieldname}' in table '{worksheet}' "
                        f"has duplicates {','.join(duplicates)}"
                    )
                    return
                if response.startswith(" "):
                    self._error(
                        f"ERROR - Responses: Invalid static radio button options for '{fieldname}' in table '{worksheet}'. "
                        "Please remove leading spaces."
                    )
                    return
                if ": " in response:
                    self._error(
                        f"ERROR - Responses: Invalid static radio button options for '{fieldname}' in table '{worksheet}'. "
                        "Please remove space after the colon (:) for static responses."
                    )
                    return

    def _check_numeric_range(self, worksheet: str, value: str, fieldname: str, range_name: str) -> None:
        if not self.DECIMAL_RE.fullmatch(value):
            self._error(
                f"ERROR - {range_name}: FieldName '{fieldname}' in worksheet '{worksheet}' "
                f"has a non-numeric value for {range_name}: {value}"
            )

    def _check_date_range(self, worksheet: str, value: str, fieldname: str, range_name: str) -> None:
        if value == "-9":
            self._error(
                f"ERROR - {range_name}: FieldName '{fieldname}' in worksheet '{worksheet}' has a missing value for {range_name}"
            )
            return
        if value in {"0", "+0d", "-0d"}:
            return
        if self.DATE_RANGE_RE.fullmatch(value):
            return
        if self.HARDCODED_DATE_RE.fullmatch(value):
            try:
                ET.fromstring(f"<d>{value}</d>")
            except Exception:
                pass
            try:
                from datetime import datetime

                datetime.strptime(value, "%Y-%m-%d")
                return
            except ValueError:
                self._error(
                    f"ERROR - {range_name}: FieldName '{fieldname}' in worksheet '{worksheet}' has an invalid date value: {value}"
                )
                return
        self._error(
            f"ERROR - {range_name}: FieldName '{fieldname}' in worksheet '{worksheet}' has an invalid format for {range_name}: {value}"
        )

    def _check_logic_check_syntax(self, worksheet: str, logic_check: str, fieldname: str) -> None:
        if ";" not in logic_check:
            self._error(
                f"ERROR - LogicCheck: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax "
                f"for LogicCheck (missing semicolon): {logic_check}"
            )
            return
        parts = logic_check.split(";", 1)
        if len(parts) != 2:
            self._error(
                f"ERROR - LogicCheck: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for LogicCheck: {logic_check}"
            )
            return
        expression = parts[0].strip()
        message = parts[1].strip()
        if not (message.startswith("'") and message.endswith("'")):
            self._error(
                f"ERROR - LogicCheck: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for LogicCheck "
                f"(message must be in single quotes): {logic_check}"
            )
            return
        operators = ["=", "!=", "<>", ">", ">=", "<", "<=", "and", "or"]
        if not any(op in expression for op in operators):
            self._error(
                f"ERROR - LogicCheck: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for LogicCheck "
                f"(no operator found): {logic_check}"
            )

    def _check_special_button(self, worksheet: str, value: str, fieldname: str, button_name: str) -> None:
        if value not in {"True", "False"}:
            self._error(
                f"ERROR: - {button_name} FieldName '{fieldname}' in worksheet '{worksheet}' "
                f"has an invalid value for '{button_name}': {value}"
            )

    def _check_skip_syntax(self, worksheet: str, skip_text: str, fieldname: str) -> None:
        skips = self._split_lines(skip_text)
        for skip in skips:
            if ":" not in skip:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return

            skip_type = "preskip" if skip[: skip.find(":")] == "preskip" else "postskip"
            if skip_type not in {"preskip", "postskip"}:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return

            parts = skip.split(",")
            if len(parts) != 2:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return

            len_skip = 13 if skip_type == "postskip" else 12
            logic_section = parts[0]
            logic_string = logic_section.split(":")
            if len(logic_string) != 2:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return

            logic_tokens = logic_section.split(" ")
            if len(logic_tokens) != 5 and "does not contain" not in logic_section:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return
            if len(logic_tokens) != 7 and "does not contain" in logic_section:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return

            space_indices = [i for i, ch in enumerate(skip) if ch == " "]
            if len(space_indices) < 4:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return
            fieldname_to_check = skip[len_skip : space_indices[2]]
            if " " in fieldname_to_check:
                self._error(f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for Skip: {skip}")
                return

            if "does not contain" not in logic_section:
                condition = skip[space_indices[2] + 1 : space_indices[3]]
                if condition not in {"=", ">", ">=", "<", "<=", "<>", "'contains'"}:
                    self._error(
                        f"ERROR - Skip: FieldName '{fieldname}' in worksheet '{worksheet}' has invalid syntax for LogicCheck: {skip}"
                    )
                    return

    def _check_logic_field_names(self, worksheet: str) -> None:
        field_index = {q.fieldName: i for i, q in enumerate(self.questionList)}
        for question in self.questionList:
            for logic_check in question.logicChecks:
                cur_field = question.fieldName
                expression = logic_check.split(";", 1)[0].strip()
                clean_expression = self.QUOTED_STRING_RE.sub("", expression)
                matches = self.FIELD_NAME_RE.findall(clean_expression)
                referenced = {m for m in matches if m.lower() not in self.LOGIC_KEYWORDS}
                for ref in referenced:
                    if ref in field_index:
                        if field_index[ref] > field_index[cur_field]:
                            self._error(
                                f"ERROR - LogicCheck: In worksheet '{worksheet}', the LogicCheck for FieldName '{cur_field}' "
                                f"uses a FieldName AFTER the current question: {ref}"
                            )
                    else:
                        self._error(
                            f"ERROR - LogicCheck: In worksheet '{worksheet}', the LogicCheck for FieldName '{cur_field}' "
                            f"uses a nonexistent FieldName: {ref}"
                        )

    def _check_skip_to_field_names(self, worksheet: str) -> None:
        field_index = {q.fieldName: i for i, q in enumerate(self.questionList)}
        for question in self.questionList:
            if not question.skip:
                continue
            cur_field = question.fieldName
            for skip in self._split_lines(question.skip):
                words = [w for w in skip.split(" ") if w]
                if len(words) < 4:
                    self._error(
                        f"ERROR - Skip: In worksheet '{worksheet}', the skip for FieldName '{cur_field}' has invalid structure: {skip}"
                    )
                    continue
                fieldname_to_check = words[2].strip().strip(",")
                fieldname_to_skip_to = words[-1].strip()
                cur_index = field_index[cur_field]

                if fieldname_to_check in field_index:
                    if field_index[fieldname_to_check] > cur_index:
                        self._error(
                            f"ERROR - Skip: In worksheet '{worksheet}', the skip for FieldName '{cur_field}' "
                            f"checks skip for a FieldName AFTER the current question: {fieldname_to_check}"
                        )
                else:
                    self._error(
                        f"ERROR - Skip: In worksheet '{worksheet}', the skip for FieldName '{cur_field}' "
                        f"checks skip of a nonexistent FieldName: {fieldname_to_check}"
                    )

                if fieldname_to_skip_to in field_index:
                    target_index = field_index[fieldname_to_skip_to]
                    if target_index < cur_index:
                        self._error(
                            f"ERROR - Skip: In worksheet '{worksheet}', the skip for FieldName '{cur_field}' "
                            f"skips to a FieldName BEFORE the current question: {fieldname_to_skip_to}"
                        )
                    elif target_index == cur_index:
                        self._error(
                            f"ERROR - Skip: In worksheet '{worksheet}', the skip for FieldName '{cur_field}' "
                            f"skips to the current question: {fieldname_to_skip_to}"
                        )
                else:
                    self._error(
                        f"ERROR - Skip: In worksheet '{worksheet}', the skip for FieldName '{cur_field}' "
                        f"skips to a nonexistent FieldName: {fieldname_to_skip_to}"
                    )

    def _check_required_max_characters(self, worksheet: str) -> None:
        for question in self.questionList:
            if (
                question.fieldType in {"text", "text_integer", "phone_num"}
                and question.questionType not in {"automatic", "checkbox", "combobox"}
                and question.maxCharacters == "-9"
            ):
                self._error(
                    f"ERROR - MaxCharacters: In worksheet '{worksheet}', MaxCharacters for FieldName '{question.fieldName}' needs a value"
                )

    def _check_duplicate_columns(self, worksheet: str) -> None:
        fields = [q.fieldName for q in self.questionList if q.questionType != "information"]
        duplicates = sorted({f for f in fields if fields.count(f) > 1})
        if len(set(fields)) != len(fields):
            self._error(
                "ERROR - Duplicate fieldnames found in worksheet: "
                f"{worksheet}. Duplicated fieldnames: {','.join(duplicates)}. "
                "Check for empty rows at the end of the spreadsheet and delete them."
            )

    def _parse_operator(self, op: str) -> str:
        op = op.strip()
        if op == ">":
            return "&gt;"
        if op == "<":
            return "&lt;"
        if op == ">=":
            return "&gt;="
        if op == "<=":
            return "&lt;="
        if op in {"=", "!=", "<>"}:
            return op
        return "="

    def _parse_dynamic_responses(self, responses: str, question: Question, worksheet: str, fieldname: str) -> None:
        for line in self._split_lines(responses):
            trimmed = line.strip()
            if not trimmed:
                continue
            parts = trimmed.split(":", 1)
            if len(parts) != 2:
                self._error(
                    f"ERROR - Responses: Invalid dynamic response line format for FieldName '{fieldname}' "
                    f"in worksheet '{worksheet}': '{trimmed}'"
                )
                continue
            key = parts[0].strip().lower()
            value = parts[1].strip()

            if key == "source":
                lowered = value.lower()
                if lowered == "csv":
                    question.responseSourceType = ResponseSourceType.CSV
                elif lowered == "database":
                    question.responseSourceType = ResponseSourceType.DATABASE
                else:
                    self._error(
                        f"ERROR - Responses: Invalid source type '{value}' for FieldName '{fieldname}' in worksheet '{worksheet}'. "
                        "Must be 'csv' or 'database'."
                    )
            elif key == "file":
                question.responseSourceFile = value
            elif key == "table":
                question.responseSourceTable = value
            elif key == "filter":
                match = self.FILTER_MATCH_RE.match(value)
                if match:
                    question.responseFilters.append(
                        Filter(
                            column=match.group(1).strip(),
                            operator=self._parse_operator(match.group(2) if match.group(2) else "="),
                            value=match.group(3).strip(),
                        )
                    )
                else:
                    self._error(
                        f"ERROR - Responses: Invalid filter format for FieldName '{fieldname}' in worksheet '{worksheet}': "
                        f"'{value}'. Expected 'column [operator] value'."
                    )
            elif key == "display":
                question.responseDisplayColumn = value
            elif key == "value":
                question.responseValueColumn = value
            elif key == "distinct":
                lowered = value.lower()
                if lowered == "true":
                    question.responseDistinct = True
                elif lowered == "false":
                    question.responseDistinct = False
                else:
                    self._error(
                        f"ERROR - Responses: Invalid boolean value for 'distinct' for FieldName '{fieldname}' in worksheet '{worksheet}'. "
                        "Must be 'true' or 'false'."
                    )
            elif key == "empty_message":
                question.responseEmptyMessage = value
            elif key == "dont_know":
                parts = value.split(",", 1)
                question.responseDontKnowValue = parts[0].strip()
                if len(parts) > 1:
                    question.responseDontKnowLabel = parts[1].strip()
            elif key == "not_in_list":
                parts = value.split(",", 1)
                question.responseNotInListValue = parts[0].strip()
                if len(parts) > 1:
                    question.responseNotInListLabel = parts[1].strip()
            else:
                self.logstring.append(
                    f"WARNING - Responses: Unknown dynamic response key '{key}' for FieldName '{fieldname}' in worksheet '{worksheet}'."
                )

    def _parse_automatic_calculation(self, responses: str, question: Question, worksheet: str, fieldname: str) -> None:
        current_calc = ""
        when_lines: list[str] = []
        part_lines: list[str] = []
        for line in self._split_lines(responses):
            trimmed = line.strip()
            if not trimmed:
                continue
            parts = trimmed.split(":", 1)
            if len(parts) != 2:
                self._error(
                    f"ERROR - Calculation: Invalid line format for FieldName '{fieldname}' in worksheet '{worksheet}': '{trimmed}'"
                )
                continue
            key = parts[0].strip().lower()
            value = parts[1].strip()

            if key == "calc":
                current_calc = value.lower()
                mapping = {
                    "query": CalculationType.QUERY,
                    "case": CalculationType.CASE,
                    "constant": CalculationType.CONSTANT,
                    "lookup": CalculationType.LOOKUP,
                    "math": CalculationType.MATH,
                    "concat": CalculationType.CONCAT,
                    "age_from_date": CalculationType.AGE_FROM_DATE,
                    "age_at_date": CalculationType.AGE_AT_DATE,
                    "date_offset": CalculationType.DATE_OFFSET,
                    "date_diff": CalculationType.DATE_DIFF,
                }
                if current_calc in mapping:
                    question.calculationType = mapping[current_calc]
                else:
                    self._error(
                        f"ERROR - Calculation: Invalid calculation type '{value}' for FieldName '{fieldname}' in worksheet '{worksheet}'. "
                        "Must be 'query', 'case', 'constant', 'lookup', 'math', 'concat', 'age_from_date', 'age_at_date', 'date_offset', or 'date_diff'."
                    )
            elif key == "sql":
                question.calculationQuerySql = value
            elif key == "param":
                self._parse_parameter(value, question, worksheet, fieldname)
            elif key == "when":
                when_lines.append(value)
            elif key == "else":
                if current_calc == "case":
                    question.calculationCaseElse = self._parse_result_value(value)
            elif key == "value":
                if current_calc in {"constant", "age_from_date", "age_at_date", "date_offset", "date_diff"}:
                    question.calculationConstantValue = value
            elif key == "field":
                if current_calc in {"lookup", "age_from_date", "age_at_date", "date_offset", "date_diff"}:
                    question.calculationLookupField = value
            elif key == "unit":
                if current_calc == "date_diff":
                    question.calculationUnit = value
            elif key == "operator":
                if current_calc == "math":
                    if value in {"+", "-", "*", "/"}:
                        question.calculationMathOperator = value
                    else:
                        self._error(
                            f"ERROR - Calculation: Invalid math operator '{value}' for FieldName '{fieldname}' in worksheet '{worksheet}'. Must be +, -, *, or /."
                        )
            elif key == "separator":
                if current_calc in {"concat", "age_at_date"}:
                    question.calculationConcatSeparator = value
            elif key == "part":
                part_lines.append(value)
            else:
                self.logstring.append(
                    f"WARNING - Calculation: Unknown calculation key '{key}' for FieldName '{fieldname}' in worksheet '{worksheet}'."
                )

        if current_calc == "case":
            for when_line in when_lines:
                self._parse_when_condition(when_line, question, worksheet, fieldname)

        if current_calc in {"math", "concat"}:
            for part_line in part_lines:
                part = self._parse_part_line(part_line, worksheet, fieldname)
                if not part:
                    continue
                if current_calc == "math":
                    question.calculationMathParts.append(part)
                else:
                    question.calculationConcatParts.append(part)

        self._validate_calculation_fields(question, worksheet, fieldname)

    def _parse_parameter(self, param_str: str, question: Question, worksheet: str, fieldname: str) -> None:
        match = self.PARAMETER_RE.match(param_str)
        if not match:
            self._error(
                f"ERROR - Calculation: Invalid parameter format '{param_str}' for FieldName '{fieldname}' in worksheet '{worksheet}'. "
                "Expected format: '@paramName = fieldName'."
            )
            return
        name = match.group(1).strip()
        if not name.startswith("@"):
            name = "@" + name
        question.calculationQueryParameters.append(
            CalculationParameter(name=name, fieldName=match.group(2).strip())
        )

    def _parse_when_condition(self, when_str: str, question: Question, worksheet: str, fieldname: str) -> None:
        match = self.WHEN_CONDITION_RE.match(when_str)
        if not match:
            self._error(
                f"ERROR - Calculation: Invalid when condition format '{when_str}' for FieldName '{fieldname}' in worksheet '{worksheet}'. "
                "Expected format: 'field operator value => result'."
            )
            return
        question.calculationCaseConditions.append(
            CaseCondition(
                field=match.group(1).strip(),
                operator=match.group(2).strip(),
                value=match.group(3).strip(),
                result=self._parse_result_value(match.group(4).strip()),
            )
        )

    @staticmethod
    def _parse_result_value(value: str) -> CalculationPart:
        return CalculationPart(type=CalculationType.CONSTANT, constantValue=value)

    def _parse_part_line(self, part_line: str, worksheet: str, fieldname: str) -> CalculationPart | None:
        words = part_line.split(" ", 1)
        if len(words) < 2:
            self._error(
                f"ERROR - Calculation: Invalid part format '{part_line}' for FieldName '{fieldname}' in worksheet '{worksheet}'. "
                "Expected 'type value'."
            )
            return None
        part_type = words[0].strip().lower()
        part_value = words[1].strip()
        if part_type == "constant":
            return CalculationPart(type=CalculationType.CONSTANT, constantValue=part_value)
        if part_type == "lookup":
            return CalculationPart(type=CalculationType.LOOKUP, lookupField=part_value)
        if part_type == "query":
            return CalculationPart(type=CalculationType.QUERY, querySql=part_value)

        self._error(
            f"ERROR - Calculation: Invalid part type '{part_type}' for FieldName '{fieldname}' in worksheet '{worksheet}'. "
            "Must be 'constant', 'lookup', or 'query'."
        )
        return None

    def _validate_calculation_fields(self, question: Question, worksheet: str, fieldname: str) -> None:
        ctype = question.calculationType
        if ctype == CalculationType.QUERY and not question.calculationQuerySql:
            self._error(
                f"ERROR - Calculation: Query calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                "is missing required 'sql' field."
            )
        elif ctype == CalculationType.CASE and len(question.calculationCaseConditions) == 0:
            self._error(
                f"ERROR - Calculation: Case calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                "is missing 'when' conditions."
            )
        elif ctype == CalculationType.CONSTANT and not question.calculationConstantValue:
            self._error(
                f"ERROR - Calculation: Constant calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                "is missing required 'value' field."
            )
        elif ctype == CalculationType.LOOKUP and not question.calculationLookupField:
            self._error(
                f"ERROR - Calculation: Lookup calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                "is missing required 'field' field."
            )
        elif ctype == CalculationType.MATH:
            if not question.calculationMathOperator:
                self._error(
                    f"ERROR - Calculation: Math calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'operator' field."
                )
            if len(question.calculationMathParts) < 2:
                self._error(
                    f"ERROR - Calculation: Math calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "must have at least 2 parts."
                )
        elif ctype == CalculationType.CONCAT and len(question.calculationConcatParts) == 0:
            self._error(
                f"ERROR - Calculation: Concat calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                "must have at least 1 part."
            )
        elif ctype == CalculationType.AGE_FROM_DATE:
            if not question.calculationLookupField:
                self._error(
                    f"ERROR - Calculation: AgeFromDate calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'field' field."
                )
            if not question.calculationConstantValue:
                self._error(
                    f"ERROR - Calculation: AgeFromDate calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'value' field."
                )
        elif ctype == CalculationType.AGE_AT_DATE:
            if not question.calculationLookupField:
                self._error(
                    f"ERROR - Calculation: AgeAtDate calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'field' field."
                )
            if not question.calculationConstantValue:
                self._error(
                    f"ERROR - Calculation: AgeAtDate calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'value' field."
                )
        elif ctype == CalculationType.DATE_OFFSET:
            if not question.calculationLookupField:
                self._error(
                    f"ERROR - Calculation: DateOffset calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'field' field."
                )
            if not question.calculationConstantValue:
                self._error(
                    f"ERROR - Calculation: DateOffset calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'value' field."
                )
            elif not self.DATE_RANGE_RE.fullmatch(question.calculationConstantValue):
                self._error(
                    f"ERROR - Calculation: DateOffset calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    f"has invalid 'value' format: {question.calculationConstantValue}. Expected format like '+28d', '-1y', etc."
                )
        elif ctype == CalculationType.DATE_DIFF:
            if not question.calculationLookupField:
                self._error(
                    f"ERROR - Calculation: DateDiff calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'field' field (start date)."
                )
            if not question.calculationConstantValue:
                self._error(
                    f"ERROR - Calculation: DateDiff calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'value' field (end date)."
                )
            if not question.calculationUnit:
                self._error(
                    f"ERROR - Calculation: DateDiff calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    "is missing required 'unit' field."
                )
            elif question.calculationUnit.lower() not in {"d", "w", "m", "y"}:
                self._error(
                    f"ERROR - Calculation: DateDiff calculation for FieldName '{fieldname}' in worksheet '{worksheet}' "
                    f"has invalid 'unit': {question.calculationUnit}. Must be 'd', 'w', 'm', or 'y'."
                )

